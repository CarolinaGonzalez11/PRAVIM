from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from .models import FichaAcogida, PersonaAtendida, Denuncia, MotivoIngreso, AspectosPsicologicos, RedesApoyo, PlanAccion, Egreso
from .forms import (
    FichaAcogidaForm,
    PersonaAtendidaForm,
    DenunciaForm,
    MotivoIngresoForm,
    AspectosPsicologicosForm,
    RedesApoyoForm,
    PlanAccionForm,
    IntervencionForm,
    EgresoForm
    )
from django.forms import modelformset_factory
from datetime import datetime
from datetime import date, timedelta
from .models import FichaChangeLog


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.urls import reverse
from .models import (
    FichaAcogida, PersonaAtendida, Denuncia, MotivoIngreso, AspectosPsicologicos,
    RedesApoyo, PlanAccion
)
from .forms import (
    FichaAcogidaForm, PersonaAtendidaForm, DenunciaForm,
    MotivoIngresoForm, AspectosPsicologicosForm, RedesApoyoForm,
    PlanAccionForm, EgresoForm,
    DerivacionCavdForm, DerivacionUravitForm, DerivacionCdmCaiForm,
    DerivacionSaludForm, DerivacionOfamForm, DerivacionDidecoForm,
    DerivacionGestionTerritorialForm, DerivacionCapsUdlaForm,
    DerivacionOlnForm, DerivacionOtroForm
)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.urls import reverse

from .models import FichaAcogida
from .forms import (
    FichaAcogidaForm, PersonaAtendidaForm, DenunciaForm, MotivoIngresoForm,
    AspectosPsicologicosForm, RedesApoyoForm, PlanAccionForm, EgresoForm,
    DerivacionCavdForm, DerivacionUravitForm, DerivacionCdmCaiForm,
    DerivacionSaludForm, DerivacionOfamForm, DerivacionDidecoForm,
    DerivacionGestionTerritorialForm, DerivacionCapsUdlaForm,
    DerivacionOlnForm, DerivacionOtroForm
)

def registrar_cambios_bitacora(instance_old, instance_new, usuario, ficha, accion=None):
    from django.forms.models import model_to_dict
    if accion == "ELIMINADO":
        # Registra la eliminación con los valores antiguos
        old = model_to_dict(instance_old) if instance_old else {}
        for campo, valor in old.items():
            FichaChangeLog.objects.create(
                ficha=ficha,
                usuario=usuario,
                campo=campo,
                valor_anterior=valor,
                valor_nuevo="—",
                accion="ELIMINADO"
            )
        return

    new = model_to_dict(instance_new) if instance_new else {}
    old = model_to_dict(instance_old) if instance_old else {}

    if not old:
        for campo, valor in new.items():
            if valor not in [None, '']:
                FichaChangeLog.objects.create(
                    ficha=ficha,
                    usuario=usuario,
                    campo=campo,
                    valor_anterior="—",
                    valor_nuevo=valor,
                    accion="CREADO"
                )
    else:
        for campo, nuevo_valor in new.items():
            antiguo_valor = old.get(campo)
            if antiguo_valor != nuevo_valor:
                FichaChangeLog.objects.create(
                    ficha=ficha,
                    usuario=usuario,
                    campo=campo,
                    valor_anterior=antiguo_valor,
                    valor_nuevo=nuevo_valor,
                    accion="EDITADO"
                )


DERIV_FORMS = [
    ('cavd', DerivacionCavdForm),
    ('uravit', DerivacionUravitForm),
    ('cdm_cai', DerivacionCdmCaiForm),
    ('salud', DerivacionSaludForm),
    ('ofam', DerivacionOfamForm),
    ('dideco', DerivacionDidecoForm),
    ('gestion_territorial', DerivacionGestionTerritorialForm),
    ('caps_udla', DerivacionCapsUdlaForm),
    ('oln', DerivacionOlnForm),
    ('otro', DerivacionOtroForm),
]

@login_required
def crear_ficha_completa(request):
    # Formularios principales
    if request.method == 'POST':
        form_ficha    = FichaAcogidaForm(request.POST, prefix='ficha')
        form_persona  = PersonaAtendidaForm(request.POST, prefix='persona')
        form_denuncia = DenunciaForm(request.POST, prefix='denuncia')
        form_motivo   = MotivoIngresoForm(request.POST, prefix='motivo')
        form_psico    = AspectosPsicologicosForm(request.POST, prefix='psico')
        form_redes    = RedesApoyoForm(request.POST, prefix='redes')
        form_plan     = PlanAccionForm(request.POST, prefix='plan')

        # Formularios de derivación
        deriv_forms = {
            key: form_cls(request.POST, prefix=f'deriv_{key}')
            for key, form_cls in DERIV_FORMS
        }

        # Validación conjunta de los principales
        principales_valid = all([
            form_ficha.is_valid(), form_persona.is_valid(), form_denuncia.is_valid(),
            form_motivo.is_valid(), form_psico.is_valid(),
            form_redes.is_valid(), form_plan.is_valid()
        ])

        if principales_valid:
            # Guardar ficha y relacionados
            ficha = form_ficha.save(commit=False)
            ficha.profesional = request.user
            ficha.save()

            # Guardar cada objeto por separado y registrar en bitácora
            persona = form_persona.save(commit=False)
            persona.ficha = ficha
            persona.save()

            denuncia = form_denuncia.save(commit=False)
            denuncia.ficha = ficha
            denuncia.save()

            motivo = form_motivo.save(commit=False)
            motivo.ficha = ficha
            motivo.save()

            psicologico = form_psico.save(commit=False)
            psicologico.ficha = ficha
            psicologico.save()

            redes = form_redes.save(commit=False)
            redes.ficha = ficha
            redes.save()

            plan = form_plan.save(commit=False)
            plan.ficha = ficha
            plan.save()


            # --- AQUI REGISTRAS EN BITÁCORA ---
            registrar_cambios_bitacora(None, ficha, request.user, ficha, accion="CREADO")
            registrar_cambios_bitacora(None, persona, request.user, ficha, accion="CREADO")
            registrar_cambios_bitacora(None, denuncia, request.user, ficha, accion="CREADO")
            registrar_cambios_bitacora(None, motivo, request.user, ficha, accion="CREADO")
            registrar_cambios_bitacora(None, psicologico, request.user, ficha, accion="CREADO")
            registrar_cambios_bitacora(None, redes, request.user, ficha, accion="CREADO")
            registrar_cambios_bitacora(None, plan, request.user, ficha, accion="CREADO")

            # Guardar solo las derivaciones con el checkbox marcado
            for key, deriv_form in deriv_forms.items():
                if request.POST.get(f'deriv_{key}-check'):
                    if deriv_form.is_valid():
                        drv = deriv_form.save(commit=False)
                        drv.ficha = ficha
                        drv.save()
                        # También registra la derivación en bitácora
                        registrar_cambios_bitacora(None, drv, request.user, ficha, accion="CREADO")

            return redirect(f'{reverse("fichas:ver_ficha_completa", args=[ficha.id])}?guardado=1')
    else:
        # GET: instanciar vacíos
        form_ficha    = FichaAcogidaForm(prefix='ficha')
        form_persona  = PersonaAtendidaForm(prefix='persona')
        form_denuncia = DenunciaForm(prefix='denuncia')
        form_motivo   = MotivoIngresoForm(prefix='motivo')
        form_psico    = AspectosPsicologicosForm(prefix='psico')
        form_redes    = RedesApoyoForm(prefix='redes')
        form_plan     = PlanAccionForm(prefix='plan')
        deriv_forms   = {
            key: form_cls(prefix=f'deriv_{key}')
            for key, form_cls in DERIV_FORMS
        }

    context = {
        **{f'form_{n}': f for n, f in [
            ('ficha', form_ficha), ('persona', form_persona),
            ('denuncia', form_denuncia), ('motivo', form_motivo),
            ('psico', form_psico), ('redes', form_redes),
            ('plan', form_plan), 
        ]},
        **{f'deriv_{k}': f for k, f in deriv_forms.items()},
        'forms_list': [form_ficha, form_persona, form_denuncia,
                       form_motivo, form_psico, form_redes,
                       form_plan] + list(deriv_forms.values()),
    }
    return render(request, 'fichas/ficha_completa.html', context)



# fichas/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.urls import reverse
from .models import (
    FichaAcogida, PersonaAtendida, Denuncia, MotivoIngreso,
    AspectosPsicologicos, RedesApoyo, PlanAccion
)
from .forms import (
    FichaAcogidaForm, PersonaAtendidaForm, DenunciaForm, MotivoIngresoForm,
    AspectosPsicologicosForm, RedesApoyoForm, PlanAccionForm
)
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages

from .models import (
    FichaAcogida,
    PersonaAtendida,
    Denuncia,
    MotivoIngreso,
    AspectosPsicologicos,
    RedesApoyo,
    PlanAccion,
    Egreso,
    Intervencion,
    # Derivaciones:
    DerivacionCavd,  # <--- Este es el nombre correcto, NO "DerivacionCAVD"
    DerivacionUravit,
    DerivacionCdmCai,
    DerivacionSalud,
    DerivacionOfam,
    DerivacionDideco,
    DerivacionGestionTerritorial,
    DerivacionCapsUdla,
    DerivacionOln,
    DerivacionOtro,
)
from django.forms import modelform_factory
from django.contrib.auth.decorators import login_required
from django.urls import reverse

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages

from .models import (
    FichaAcogida, PersonaAtendida, Denuncia, MotivoIngreso, AspectosPsicologicos,
    RedesApoyo, PlanAccion,
    DerivacionCavd, DerivacionUravit, DerivacionCdmCai, DerivacionSalud,
    DerivacionOfam, DerivacionDideco, DerivacionGestionTerritorial,
    DerivacionCapsUdla, DerivacionOln, DerivacionOtro,
)

from .forms import (
    FichaAcogidaForm, PersonaAtendidaForm, DenunciaForm, MotivoIngresoForm,
    AspectosPsicologicosForm, RedesApoyoForm, PlanAccionForm,
    DerivacionCavdForm, DerivacionUravitForm, DerivacionCdmCaiForm, DerivacionSaludForm,
    DerivacionOfamForm, DerivacionDidecoForm, DerivacionGestionTerritorialForm,
    DerivacionCapsUdlaForm, DerivacionOlnForm, DerivacionOtroForm,
)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import (
    FichaAcogida, PersonaAtendida, Denuncia, MotivoIngreso, AspectosPsicologicos,
    RedesApoyo, PlanAccion,
    DerivacionCavd, DerivacionUravit, DerivacionCdmCai, DerivacionSalud,
    DerivacionOfam, DerivacionDideco, DerivacionGestionTerritorial,
    DerivacionCapsUdla, DerivacionOln, DerivacionOtro,
)
from .forms import (
    FichaAcogidaForm, PersonaAtendidaForm, DenunciaForm, MotivoIngresoForm,
    AspectosPsicologicosForm, RedesApoyoForm, PlanAccionForm,
    DerivacionCavdForm, DerivacionUravitForm, DerivacionCdmCaiForm,
    DerivacionSaludForm, DerivacionOfamForm, DerivacionDidecoForm,
    DerivacionGestionTerritorialForm, DerivacionCapsUdlaForm,
    DerivacionOlnForm, DerivacionOtroForm,
)

# Listado DRY de tuplas (nombre_campo, modelo, form, prefix)
DERIVS = [
    ('deriv_cavd', DerivacionCavd, DerivacionCavdForm, 'deriv_cavd'),
    ('deriv_uravit', DerivacionUravit, DerivacionUravitForm, 'deriv_uravit'),
    ('deriv_cdm_cai', DerivacionCdmCai, DerivacionCdmCaiForm, 'deriv_cdm_cai'),
    ('deriv_salud', DerivacionSalud, DerivacionSaludForm, 'deriv_salud'),
    ('deriv_ofam', DerivacionOfam, DerivacionOfamForm, 'deriv_ofam'),
    ('deriv_dideco', DerivacionDideco, DerivacionDidecoForm, 'deriv_dideco'),
    ('deriv_gestion_territorial', DerivacionGestionTerritorial, DerivacionGestionTerritorialForm, 'deriv_gestion_territorial'),
    ('deriv_caps_udla', DerivacionCapsUdla, DerivacionCapsUdlaForm, 'deriv_caps_udla'),
    ('deriv_oln', DerivacionOln, DerivacionOlnForm, 'deriv_oln'),
    ('deriv_otro', DerivacionOtro, DerivacionOtroForm, 'deriv_otro'),
]

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages

# ... tus imports de modelos y forms y DERIVS ...

@login_required
def editar_ficha(request, ficha_id):
    ficha = get_object_or_404(FichaAcogida, id=ficha_id)
    persona = getattr(ficha, "personaatendida", None)
    denuncia = getattr(ficha, "denuncia", None)
    motivo = getattr(ficha, "motivoingreso", None)
    psicologico = getattr(ficha, "aspectospsicologicos", None)
    redes = getattr(ficha, "redesapoyo", None)
    plan = getattr(ficha, "planaccion", None)

    derivs_instances = {
        campo: modelo.objects.filter(ficha=ficha).first()
        for campo, modelo, _, _ in DERIVS
    }

    if request.method == 'POST':
        form_ficha = FichaAcogidaForm(request.POST, instance=ficha, prefix='ficha')
        form_persona = PersonaAtendidaForm(request.POST, instance=persona, prefix='persona')
        form_denuncia = DenunciaForm(request.POST, instance=denuncia, prefix='denuncia')
        form_motivo = MotivoIngresoForm(request.POST, instance=motivo, prefix='motivo')
        form_psicologico = AspectosPsicologicosForm(request.POST, instance=psicologico, prefix='psico')
        form_redes = RedesApoyoForm(request.POST, instance=redes, prefix='redes')
        form_plan = PlanAccionForm(request.POST, instance=plan, prefix='plan')

        derivs_forms = {
            f'form_{campo}': form_class(request.POST, instance=derivs_instances[campo], prefix=prefix)
            for campo, _, form_class, prefix in DERIVS
        }

        all_valid = (
            form_ficha.is_valid() and form_persona.is_valid() and form_denuncia.is_valid()
            and form_motivo.is_valid() and form_psicologico.is_valid() and form_redes.is_valid()
            and form_plan.is_valid()
            and all(df.is_valid() for df in derivs_forms.values())
        )

        if all_valid:
            ficha_old = FichaAcogida.objects.get(id=ficha.id)
            form_ficha.save()
            ficha.refresh_from_db()
            registrar_cambios_bitacora(ficha_old, ficha, request.user, ficha)

            for form, obj, rel, modelo in [
                (form_persona, persona, 'personaatendida', PersonaAtendida),
                (form_denuncia, denuncia, 'denuncia', Denuncia),
                (form_motivo, motivo, 'motivoingreso', MotivoIngreso),
                (form_psicologico, psicologico, 'aspectospsicologicos', AspectosPsicologicos),
                (form_redes, redes, 'redesapoyo', RedesApoyo),
                (form_plan, plan, 'planaccion', PlanAccion),
            ]:
                if obj:
                    old_obj = modelo.objects.get(id=obj.id)
                else:
                    old_obj = None
                instance = form.save(commit=False)
                instance.ficha = ficha
                instance.save()
                registrar_cambios_bitacora(old_obj, instance, request.user, ficha)

            for campo, modelo, _, prefix in DERIVS:
                checkbox_name = f"{campo}-check"
                dform = derivs_forms[f'form_{campo}']
                deriv = derivs_instances[campo]
                if request.POST.get(checkbox_name):
                    if deriv:
                        old_deriv = modelo.objects.get(id=deriv.id)
                    else:
                        old_deriv = None
                    instance = dform.save(commit=False)
                    instance.ficha = ficha
                    instance.save()
                    registrar_cambios_bitacora(old_deriv, instance, request.user, ficha)
                else:
                    if deriv:
                        deriv.delete()

                        
                        

            messages.success(request, "Ficha actualizada correctamente.")
            return redirect('fichas:ver_ficha_completa', ficha.id)
    else:
        form_ficha = FichaAcogidaForm(instance=ficha, prefix='ficha')
        form_persona = PersonaAtendidaForm(instance=persona, prefix='persona')
        form_denuncia = DenunciaForm(instance=denuncia, prefix='denuncia')
        form_motivo = MotivoIngresoForm(instance=motivo, prefix='motivo')
        form_psicologico = AspectosPsicologicosForm(instance=psicologico, prefix='psico')
        form_redes = RedesApoyoForm(instance=redes, prefix='redes')
        form_plan = PlanAccionForm(instance=plan, prefix='plan')

        
        derivs_forms = {
            f'form_{campo}': form_class(instance=derivs_instances[campo], prefix=prefix)
            for campo, _, form_class, prefix in DERIVS
        }

    context = {
        'ficha': ficha,
        'form_ficha': form_ficha,
        'form_persona': form_persona,
        'form_denuncia': form_denuncia,
        'form_motivo': form_motivo,
        'form_psicologico': form_psicologico,
        'form_redes': form_redes,
        'form_plan': form_plan,
        **derivs_forms,
    }
    return render(request, 'fichas/editar_ficha.html', context)



from .models import (
    FichaAcogida,
    DerivacionCavd,
    DerivacionUravit,
    DerivacionCdmCai,
    DerivacionSalud,
    DerivacionOfam,
    DerivacionDideco,
    DerivacionGestionTerritorial,
    DerivacionCapsUdla,
    DerivacionOln,
    DerivacionOtro,
)
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, render

from .models import (
    FichaAcogida,
    PersonaAtendida,
    Denuncia,
    MotivoIngreso,
    AspectosPsicologicos,
    RedesApoyo,
    PlanAccion,
    Egreso,
    Intervencion,
    DerivacionCavd,
    DerivacionUravit,
    DerivacionCdmCai,
    DerivacionSalud,
    DerivacionOfam,
    DerivacionDideco,
    DerivacionGestionTerritorial,
    DerivacionCapsUdla,
    DerivacionOln,
    DerivacionOtro,
)


@login_required
def ver_ficha_completa(request, ficha_id):
    # Obtengo la ficha o 404
    ficha = get_object_or_404(FichaAcogida, id=ficha_id)

    # Relaciones uno-a-uno
    persona     = getattr(ficha, 'personaatendida', None)
    denuncia    = getattr(ficha, 'denuncia', None)
    motivo      = getattr(ficha, 'motivoingreso', None)
    psicologico = getattr(ficha, 'aspectospsicologicos', None)
    redes       = getattr(ficha, 'redesapoyo', None)
    plan        = getattr(ficha, 'planaccion', None)
    egreso      = getattr(ficha, 'egreso', None)

    # Todas las intervenciones vinculadas, ordenadas por fecha
    intervenciones = Intervencion.objects.filter(ficha=ficha).order_by('fecha')

    # Derivaciones (OneToOneField en cada modelo)
    deriv_cavd               = getattr(ficha, 'derivacion_cavd', None)
    deriv_uravit             = getattr(ficha, 'derivacion_uravit', None)
    deriv_cdm_cai            = getattr(ficha, 'derivacion_cdm_cai', None)
    deriv_salud              = getattr(ficha, 'derivacion_salud', None)
    deriv_ofam               = getattr(ficha, 'derivacion_ofam', None)
    deriv_dideco             = getattr(ficha, 'derivacion_dideco', None)
    deriv_gestion_territorial= getattr(ficha, 'derivacion_gestion_territorial', None)
    deriv_caps_udla          = getattr(ficha, 'derivacion_caps_udla', None)
    deriv_oln                = getattr(ficha, 'derivacion_oln', None)
    deriv_otro               = getattr(ficha, 'derivacion_otro', None)

    context = {
        'ficha': ficha,
        'persona': persona,
        'denuncia': denuncia,
        'motivo': motivo,
        'psicologico': psicologico,
        'redes': redes,
        'plan': plan,
        'egreso': egreso,
        'intervenciones': intervenciones,
        # Paso cada derivación para que en el template puedas hacer {% if deriv_cavd %} … {% endif %}
        'deriv_cavd': deriv_cavd,
        'deriv_uravit': deriv_uravit,
        'deriv_cdm_cai': deriv_cdm_cai,
        'deriv_salud': deriv_salud,
        'deriv_ofam': deriv_ofam,
        'deriv_dideco': deriv_dideco,
        'deriv_gestion_territorial': deriv_gestion_territorial,
        'deriv_caps_udla': deriv_caps_udla,
        'deriv_oln': deriv_oln,
        'deriv_otro': deriv_otro,
    }
    return render(request, 'fichas/ver_ficha.html', context)


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import FichaAcogida, PersonaAtendida, Denuncia, MotivoIngreso, AspectosPsicologicos, RedesApoyo, PlanAccion
from .forms import (
    FichaAcogidaForm, PersonaAtendidaForm, DenunciaForm, MotivoIngresoForm,
    AspectosPsicologicosForm, RedesApoyoForm, PlanAccionForm
)

from django.shortcuts import render
from django.db.models import Q
from django.core.paginator import Paginator

from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from fichas.models import FichaAcogida

@login_required
def buscar_persona(request):
    fichas = FichaAcogida.objects.all().select_related('personaatendida', 'motivoingreso', 'egreso')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    numero_ficha = request.GET.get('numero_ficha')
    nombre = request.GET.get('nombre')
    rut = request.GET.get('rut')
    domicilio = request.GET.get('domicilio')
    
    delito = request.GET.get('delito')
    estado = request.GET.get('estado')
    ingreso = request.GET.get('ingreso')
    if ingreso == "con_ingreso":
        fichas = fichas.filter(ingreso_efectivo=True)
    elif ingreso == "sin_ingreso":
        fichas = fichas.filter(ingreso_efectivo=False)

    if fecha_inicio:
        fichas = fichas.filter(fecha_recepcion__gte=fecha_inicio)
    if fecha_fin:
        fichas = fichas.filter(fecha_recepcion__lte=fecha_fin)
    if numero_ficha:
        fichas = fichas.filter(id=numero_ficha)
    if nombre:
        fichas = fichas.filter(personaatendida__nombre__icontains=nombre)
    if delito:
        fichas = fichas.filter(motivoingreso__motivo_ingreso=delito)
    if estado == "ABIERTA":
        fichas = fichas.filter(egreso__isnull=True)
    elif estado == "CERRADA":
        fichas = fichas.filter(egreso__isnull=False)
    if rut:
        fichas = fichas.filter(personaatendida__rut_pasaporte__icontains=rut)
    if domicilio:
        fichas = fichas.filter(personaatendida__direccion__icontains=domicilio)

    # ORDENAMIENTO
    ordenar_por = request.GET.get('ordenar_por', 'id')
    sentido = request.GET.get('sentido', 'desc')
    if not sentido:
        sentido = 'desc' 
    
    campos_permitidos = {
        'id': 'id',
        'fecha_recepcion': 'fecha_recepcion',
        'nombre': 'personaatendida__nombre',
        'delito': 'motivoingreso__motivo_ingreso',
    }
    campo = campos_permitidos.get(ordenar_por, 'id')
    if sentido == 'desc':
        campo = '-' + campo
    fichas = fichas.order_by(campo)
    hoy = date.today()
    for ficha in fichas:
        ultima = ficha.intervencion_set.order_by('-fecha').first()
        ficha.fecha_ultima_intervencion = ultima.fecha if ultima else None
        ficha.destacar_intervencion = (
            ultima and (hoy - ultima.fecha).days > 14
        )

    # PAGINACIÓN
    paginator = Paginator(fichas, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Agrega esto para pasar las opciones del select
    OPCIONES_MOTIVO = getattr(FichaAcogida, 'OPCIONES_MOTIVO', [
        ('FEMICIDIO', 'Femicidio'),
        ('PARRICIDIO', 'Parricidio'),
        ('SECUESTRO', 'Secuestro / Intento Secuestro'),
        ('DELITO_SEXUAL', 'Delito Sexual'),
        ('ROBO', 'Robo'),
        ('VIF', 'Violencia Intrafamiliar (VIF)'),
        ('SUSTRACCION', 'Sustracción de Menores'),
        ('HOMICIDIO', 'Homicidio / Cuasidelito de Homicidio'),
        ('BALACERAS', 'Balas Locas / Balaceras'),
        ('LESIONES', 'Lesiones'),
        ('ODIO', 'Crímenes de Odio'),
        ('AMENAZAS', 'Amenazas'),
        ('DAÑOS', 'Daños'),
        ('FALLECIMIENTO', 'Fallecimientos'),
        ('OTROS', 'Otros'),
    ])
    columnas = [
        ('id', 'N° Ficha'),
        ('fecha_recepcion', 'Fecha Recepción'),
        ('nombre', 'Nombre'),
        ('delito', 'Motivo/Delito'),
        ('id', 'N° Ficha'),
        ('fecha_recepcion', 'Fecha Recepción'),
        ('nombre', 'Nombre'),
        ('delito', 'Motivo/Delito'),
        ('estado', 'Estado'),
        ('ultima_intervencion', 'Últ. intervención'),        
      
    ]
    context = {
        'page_obj': page_obj,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
        'numero_ficha': numero_ficha,
        'nombre': nombre,
        'delito': delito,
        'ordenar_por': ordenar_por,
        'sentido': sentido,
        'request': request,
        'OPCIONES_MOTIVO': OPCIONES_MOTIVO,
        'columnas': columnas, # <-- esto es lo nuevo
        'estado': estado,
        'rut': rut,
        'domicilio': domicilio,
        'ingreso': ingreso,

    }
    return render(request, 'fichas/buscar_persona.html', context)

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .models import FichaAcogida, Intervencion
from .forms import IntervencionForm

@login_required
def ingresar_intervencion(request):
    ficha_id = request.GET.get('ficha_id')
    ficha = None
    intervenciones = []
    if ficha_id:
        ficha = get_object_or_404(FichaAcogida, pk=ficha_id)
        # Trae todas las intervenciones asociadas, ordenadas por fecha descendente
        intervenciones = Intervencion.objects.filter(ficha=ficha).order_by('-fecha')
    else:
        # Si por error no hay ficha, puedes redirigir a otra página o mostrar error
        return redirect('fichas:buscar_persona')
    
    # El formulario solo lo necesitas si quieres mostrar un form tradicional (no necesario si solo usas AJAX)
    form = IntervencionForm()

    columnas = [
        ('fecha', 'Fecha'),
        ('etapa', 'Etapa'),
        ('tipo_intervencion', 'Tipo de Atención'),
        ('responsables', 'Responsables/es'),
        ('participantes', 'Participantes'),
        ('lugar_via', 'Lugar o Vía de Intervención'),
        ('objetivos', 'Objetivo/s'),
        ('descripcion', 'Descripción'),
        ('resultados', 'Resultados/Sugerencias'),
    ]

    context = {
        'form': form,
        'ficha': ficha,
        'intervenciones': intervenciones,
        'columnas': columnas,  # Por si lo quieres usar en el template
    }
    return render(request, 'fichas/ingresar_intervencion.html', context)


import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from .models import Intervencion, FichaAcogida

import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required

from datetime import datetime

@csrf_exempt
@login_required
def api_intervencion_guardar(request, ficha_id):
    if request.method == "POST":
        try:
            data = json.loads(request.body.decode())
            ficha = FichaAcogida.objects.get(id=ficha_id)

            fecha_str = data.get("fecha")
            fecha_obj = None
            if fecha_str:
                try:
                    fecha_obj = datetime.strptime(fecha_str, "%Y-%m-%d").date()
                except ValueError:
                    return JsonResponse({"success": False, "error": "Formato de fecha inválido"})

            interv = Intervencion.objects.create(
                ficha=ficha,
                fecha=fecha_obj,
                etapa=data.get("etapa"),
                tipo_intervencion=data.get("tipo_intervencion"),
                responsables=data.get("responsables", ""),
                participantes=data.get("participantes", ""),
                lugar_via=data.get("lugar_via"),
                objetivos=data.get("objetivos", ""),
                descripcion=data.get("descripcion", ""),
                resultados=data.get("resultados", ""),
                profesional=request.user,
            )
            # REGISTRAR EN BITÁCORA
            registrar_cambios_bitacora(None, interv, request.user, ficha, accion="CREADO")
            
            return JsonResponse({
                "success": True,
                "id": interv.id,
                "fecha": interv.fecha.strftime("%Y-%m-%d") if interv.fecha else "",
                "etapa": interv.get_etapa_display(),
                "tipo_intervencion": interv.get_tipo_intervencion_display(),
                "responsables": interv.responsables,
                "participantes": interv.participantes,
                "lugar_via": interv.get_lugar_via_display(),
                "objetivos": interv.objetivos,
                "descripcion": interv.descripcion,
                "resultados": interv.resultados,

            })
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Método no permitido"})


@csrf_exempt
@login_required
def api_intervencion_editar(request, intervencion_id):
    if request.method == "POST":
        try:
            data = json.loads(request.body.decode())
            interv = Intervencion.objects.get(id=intervencion_id)
            old_interv = Intervencion.objects.get(id=intervencion_id)
            # Clona los datos actuales para comparar
            old_instance = Intervencion.objects.get(pk=intervencion_id)
            # Actualiza campos
            fecha_str = data.get("fecha")
            if fecha_str:
                try:
                    interv.fecha = datetime.strptime(fecha_str, "%Y-%m-%d").date()
                except ValueError:
                    return JsonResponse({"success": False, "error": "Formato de fecha inválido"})

            interv.etapa = data.get("etapa")
            interv.tipo_intervencion = data.get("tipo_intervencion")
            interv.responsables = data.get("responsables", "")
            interv.participantes = data.get("participantes", "")
            interv.lugar_via = data.get("lugar_via")
            interv.objetivos = data.get("objetivos", "")
            interv.descripcion = data.get("descripcion", "")
            interv.resultados = data.get("resultados", "")
            interv.save()
            # REGISTRAR EN BITÁCORA
            registrar_cambios_bitacora(old_instance, interv, request.user, interv.ficha, accion="EDITADO")            
            return JsonResponse({
                "success": True,
                "id": interv.id,
                "fecha": interv.fecha.strftime("%Y-%m-%d") if interv.fecha else "",
                "etapa": interv.get_etapa_display(),
                "tipo_intervencion": interv.get_tipo_intervencion_display(),
                "responsables": interv.responsables,
                "participantes": interv.participantes,
                "lugar_via": interv.get_lugar_via_display(),
                "objetivos": interv.objetivos,
                "descripcion": interv.descripcion,
                "resultados": interv.resultados,

            })
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Método no permitido"})

@login_required
def gestionar_egreso(request, ficha_id):
    ficha = get_object_or_404(FichaAcogida, id=ficha_id)
    # Busca el egreso existente o crea uno nuevo (sin guardar todavía)
    egreso = Egreso.objects.filter(ficha=ficha).first()
    if request.method == "POST":
        old_egreso = egreso if egreso else None
        form = EgresoForm(request.POST, instance=egreso)
        if form.is_valid():
            nuevo_egreso = form.save(commit=False)
            nuevo_egreso.ficha = ficha   # <-- ASÍ asignas la ficha antes de guardar
            nuevo_egreso.save()
            registrar_cambios_bitacora(old_egreso, nuevo_egreso, request.user, ficha, accion="EDITADO" if old_egreso else "CREADO")

            return redirect('fichas:ver_ficha_completa', ficha.id)
    else:
        form = EgresoForm(instance=egreso)
    return render(request, 'fichas/gestionar_egreso.html', {
        'ficha': ficha,
        'form': form,
        'egreso': egreso,
    })

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from .models import FichaAcogida, MotivoIngreso, PersonaAtendida, Intervencion, Barrio, Cuadrante
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Avg, Q, F, ExpressionWrapper, DurationField
from django.db.models.functions import TruncMonth
from .models import (
    FichaAcogida, PersonaAtendida, Denuncia, MotivoIngreso, AspectosPsicologicos,
    RedesApoyo, PlanAccion, Intervencion, Egreso, Barrio, Cuadrante,
    DerivacionSalud, DerivacionCavd, DerivacionOfam, DerivacionDideco, DerivacionGestionTerritorial,
    DerivacionCapsUdla, DerivacionOln, DerivacionOtro, DerivacionUravit, DerivacionCdmCai
)
from django.db.models import Case, When, Value, CharField, Q

@login_required
def dashboard(request):
    GET = request.GET
    barrio_id = GET.get('barrio')
    cuadrante_id = GET.get('cuadrante')
    genero = GET.get('genero')
    motivo = GET.get('motivo_ingreso')
    profesional_id = GET.get('profesional')
    estado = GET.get('estado')
    etapa = GET.get('etapa')
    fecha_inicio = GET.get('fecha_inicio')
    fecha_fin = GET.get('fecha_fin')

    # 1. QUERYS BASE (sin filtro de estado ni etapa)
    fichas_base = FichaAcogida.objects.select_related('personaatendida', 'motivoingreso', 'egreso', 'profesional')
    if barrio_id:
        fichas_base = fichas_base.filter(personaatendida__barrio_id=barrio_id)
    if cuadrante_id:
        fichas_base = fichas_base.filter(personaatendida__cuadrante_id=cuadrante_id)
    if genero:
        fichas_base = fichas_base.filter(personaatendida__genero=genero)
    if motivo:
        fichas_base = fichas_base.filter(motivoingreso__motivo_ingreso=motivo)
    if profesional_id == 'nathaly':
        fichas_base = fichas_base.filter(planaccion__profesional_nathaly=True)
    elif profesional_id == 'paloma':
        fichas_base = fichas_base.filter(planaccion__profesional_paloma=True)
    elif profesional_id == 'gloria':
        fichas_base = fichas_base.filter(planaccion__cordinadora_gloria=True)


    if fecha_inicio:
        fichas_base = fichas_base.filter(fecha_recepcion__gte=fecha_inicio)
    if fecha_fin:
        fichas_base = fichas_base.filter(fecha_recepcion__lte=fecha_fin)

    # 2. KPIs globales (sobre fichas_base)
    total_fichas = fichas_base.count()
    total_preingresos = fichas_base.filter(ingreso_efectivo=False).count()
    total_ingresos = fichas_base.filter(ingreso_efectivo=True).count()
    total_egresos = fichas_base.filter(egreso__isnull=False).count()
    total_intervenciones = Intervencion.objects.filter(ficha__in=fichas_base).count()
    casos_abiertos = fichas_base.filter(egreso__fecha_egreso__isnull=True).count()
    casos_cerrados = fichas_base.filter(egreso__fecha_egreso__isnull=False).count()
    porcentaje_ingreso_efectivo = round((total_ingresos / total_fichas) * 100, 2) if total_fichas else 0

    # 3. APLICA FILTROS COMPLETOS (incluyendo estado/etapa) para breakdowns y tablas
    fichas = fichas_base
    if estado == 'ABIERTA':
        fichas = fichas.filter(egreso__fecha_egreso__isnull=True)
    elif estado == 'CERRADA':
        fichas = fichas.filter(egreso__fecha_egreso__isnull=False)
    if etapa == "preingreso":
        fichas = fichas.filter(ingreso_efectivo=False)
    elif etapa == "ingreso":
        fichas = fichas.filter(ingreso_efectivo=True)
    elif etapa == "egreso":
        fichas = fichas.filter(egreso__isnull=False)
    
    
    total_fichas = fichas.count()
    total_preingresos = fichas.filter(ingreso_efectivo=False).count()
    total_ingresos = fichas.filter(ingreso_efectivo=True).count()
    total_egresos = fichas.filter(egreso__isnull=False).count()
    total_intervenciones = Intervencion.objects.filter(ficha__in=fichas).count()
    casos_abiertos = fichas.filter(egreso__fecha_egreso__isnull=True).count()
    casos_cerrados = fichas.filter(egreso__fecha_egreso__isnull=False).count()

    # KPI destacado
    porcentaje_ingreso_efectivo = round((total_ingresos / total_fichas) * 100, 2) if total_fichas else 0

    # Tiempo promedio ingreso-egreso (solo para fichas con ingreso y egreso)
    fichas_con_egreso = fichas.filter(
        ingreso_efectivo=True,
        egreso__isnull=False,
        fecha_ingreso__isnull=False,
        egreso__fecha_egreso__isnull=False
    ).annotate(
        duracion=ExpressionWrapper(F('egreso__fecha_egreso') - F('fecha_ingreso'), output_field=DurationField())
    )
    promedio_duracion = None
    if fichas_con_egreso.exists():
        promedio = fichas_con_egreso.aggregate(prom=Avg('duracion'))['prom']
        promedio_duracion = promedio.days if promedio else None

    # Breakdown por distribución (ejemplo género, edad/rango, nacionalidad, previsión, barrio/cuadrante, motivo ingreso, profesional)
    # Género
    genero_qs = fichas.values('personaatendida__genero').annotate(cant=Count('id')).order_by('-cant')
    genero_choices = dict(PersonaAtendida._meta.get_field('genero').choices)
    genero_labels = [genero_choices.get(g["personaatendida__genero"], "No info") if g["personaatendida__genero"] else "No info" for g in genero_qs]
    genero_data = [g["cant"] for g in genero_qs]
    # Edad/rango
    rango_qs = fichas.values('personaatendida__rango_etareo').annotate(cant=Count('id')).order_by('-cant')
    rango_choices = dict(PersonaAtendida._meta.get_field('rango_etareo').choices)
    rango_labels = [rango_choices.get(r["personaatendida__rango_etareo"], "No info") if r["personaatendida__rango_etareo"] else "No info" for r in rango_qs]
    rango_data = [r["cant"] for r in rango_qs]
    # Nacionalidad
    nacionalidad_qs = fichas.values('personaatendida__nacionalidad').annotate(cant=Count('id')).order_by('-cant')
    nacionalidad_labels = [n["personaatendida__nacionalidad"] or "No info" for n in nacionalidad_qs]
    nacionalidad_data = [n["cant"] for n in nacionalidad_qs]
    # Previsión
    prevision_qs = fichas.values('personaatendida__prevision').annotate(cant=Count('id')).order_by('-cant')
    prevision_choices = dict(PersonaAtendida._meta.get_field('prevision').choices)
    prevision_labels = [prevision_choices.get(p["personaatendida__prevision"], "No info") if p["personaatendida__prevision"] else "No info" for p in prevision_qs]
    prevision_data = [p["cant"] for p in prevision_qs]
    # Ocupación (puedes hacer tabla o gráfico, si hay muchos valores mejor tabla)
    ocupacion_qs = fichas.values('personaatendida__ocupacion').annotate(cant=Count('id')).order_by('-cant')
    ocupacion_labels = [o["personaatendida__ocupacion"] or "No info" for o in ocupacion_qs]
    ocupacion_data = [o["cant"] for o in ocupacion_qs]
    # Barrio/cuadrante
    barrio_qs = fichas.values('personaatendida__barrio__nombre').annotate(cant=Count('id')).order_by('-cant')
    barrio_labels = [b["personaatendida__barrio__nombre"] or "Sin barrio" for b in barrio_qs]
    barrio_data = [b["cant"] for b in barrio_qs]
    cuadrante_qs = fichas.values('personaatendida__cuadrante__nombre').annotate(cant=Count('id')).order_by('-cant')
    cuadrante_labels = [c["personaatendida__cuadrante__nombre"] or "Sin cuadrante" for c in cuadrante_qs]
    cuadrante_data = [c["cant"] for c in cuadrante_qs]
    # Motivo de ingreso
    motivo_qs = fichas.values('motivoingreso__motivo_ingreso').annotate(cant=Count('id')).order_by('-cant')
    motivo_choices = dict(MotivoIngreso._meta.get_field('motivo_ingreso').choices)
    motivo_labels = [motivo_choices.get(m["motivoingreso__motivo_ingreso"], m["motivoingreso__motivo_ingreso"]) for m in motivo_qs]
    motivo_data = [m["cant"] for m in motivo_qs]

    # --- Profesional a cargo (solo tres nombres posibles o Sin info) ---
    PROFESIONALES_FIJOS = [
        ("Nathaly Reyes", "nathaly"),
        ("Paloma Quinteros", "paloma"),
        ("Gloria Rivera", "gloria"),
        ("Sin información", "sin_info"),
    ]

    def get_nombre_profesional(first_name, last_name):
        nombre = (first_name or "").strip()
        apellido = (last_name or "").strip()
        completo = f"{nombre} {apellido}".strip()
        if completo.lower().startswith("nathaly"):
            return "Nathaly Reyes"
        elif completo.lower().startswith("paloma"):
            return "Paloma Quinteros"
        elif completo.lower().startswith("gloria"):
            return "Gloria Rivera"
        return "Sin información"

    # Aquí está la lógica mágica:
    profesional_contador = {"Nathaly Reyes": 0, "Paloma Quinteros": 0, "Gloria Rivera": 0, "Sin información": 0}

    if profesional_id == 'nathaly':
        # Solo cuentas los casos de Nathaly, el resto en 0
        profesional_contador["Nathaly Reyes"] = fichas.count()
    elif profesional_id == 'paloma':
        profesional_contador["Paloma Quinteros"] = fichas.count()
    elif profesional_id == 'gloria':
        profesional_contador["Gloria Rivera"] = fichas.count()
    else:
        # Si no hay filtro, calcula normalmente
        profesional_qs = fichas.values('profesional__first_name', 'profesional__last_name').annotate(cant=Count('id'))
        profesional_contador = {"Nathaly Reyes": 0, "Paloma Quinteros": 0, "Gloria Rivera": 0, "Sin información": 0}

        for p in profesional_qs:
            nombre = get_nombre_profesional(p["profesional__first_name"], p["profesional__last_name"])
            print("DEBUG:", p["profesional__first_name"], p["profesional__last_name"], "->", nombre)

            if nombre in profesional_contador:
                profesional_contador[nombre] += p["cant"]
            else:
                profesional_contador["Sin información"] += p["cant"]


    profesional_labels = [nombre for nombre, _ in PROFESIONALES_FIJOS]
    profesional_data = [profesional_contador[nombre] for nombre, _ in PROFESIONALES_FIJOS]

    
    # Diagnóstico/Salud mental (puedes sumar breakdown por diagnostico, internaciones, atenciones, etc.)
    diagnostico_qs = fichas.values('aspectospsicologicos__diagnostico').annotate(cant=Count('id')).order_by('-cant')
    diagnostico_choices = dict(AspectosPsicologicos._meta.get_field('diagnostico').choices)
    diagnostico_labels = [diagnostico_choices.get(d["aspectospsicologicos__diagnostico"], d["aspectospsicologicos__diagnostico"]) for d in diagnostico_qs]
    diagnostico_data = [d["cant"] for d in diagnostico_qs]
    # Derivaciones (por tipo)
    derivaciones = {
        "salud": DerivacionSalud.objects.filter(ficha__in=fichas).count(),
        "cavd": DerivacionCavd.objects.filter(ficha__in=fichas).count(),
        "uravit": DerivacionUravit.objects.filter(ficha__in=fichas).count(),
        "ofam": DerivacionOfam.objects.filter(ficha__in=fichas).count(),
        "dideco": DerivacionDideco.objects.filter(ficha__in=fichas).count(),
        "gestion_territorial": DerivacionGestionTerritorial.objects.filter(ficha__in=fichas).count(),
        "caps_udla": DerivacionCapsUdla.objects.filter(ficha__in=fichas).count(),
        "oln": DerivacionOln.objects.filter(ficha__in=fichas).count(),
        "cdm_cai": DerivacionCdmCai.objects.filter(ficha__in=fichas).count(),
        "otro": DerivacionOtro.objects.filter(ficha__in=fichas).count(),
    }
    total_derivaciones = sum(derivaciones.values())

    # Evolución temporal (por mes)
    fichas_mes_qs = fichas.filter(fecha_ingreso__isnull=False).annotate(mes=TruncMonth('fecha_ingreso')).values('mes').annotate(cant=Count('id')).order_by('mes')
    fichas_mes_labels = [f["mes"].strftime('%Y-%m') if f["mes"] else "Sin mes" for f in fichas_mes_qs]
    fichas_mes_data = [f["cant"] for f in fichas_mes_qs]
    egresos_mes_qs = Egreso.objects.filter(ficha__in=fichas, fecha_egreso__isnull=False).annotate(mes=TruncMonth('fecha_egreso')).values('mes').annotate(cant=Count('id')).order_by('mes')
    egresos_mes_labels = [e["mes"].strftime('%Y-%m') if e["mes"] else "Sin mes" for e in egresos_mes_qs]
    egresos_mes_data = [e["cant"] for e in egresos_mes_qs]
    intervenciones_mes_qs = Intervencion.objects.filter(ficha__in=fichas, fecha__isnull=False).annotate(mes=TruncMonth('fecha')).values('mes').annotate(cant=Count('id')).order_by('mes')
    intervenciones_mes_labels = [i["mes"].strftime('%Y-%m') if i["mes"] else "Sin mes" for i in intervenciones_mes_qs]
    intervenciones_mes_data = [i["cant"] for i in intervenciones_mes_qs]

    # Filtros para selects (deja solo lo que usas en los selects)
    barrios = Barrio.objects.all()
    cuadrantes = Cuadrante.objects.all()
    motivos = MotivoIngreso._meta.get_field('motivo_ingreso').choices
    generos = PersonaAtendida._meta.get_field('genero').choices
    profesionales = FichaAcogida.objects.values('profesional__id', 'profesional__first_name', 'profesional__last_name').distinct()
    estados = FichaAcogida.ESTADOS
    etapas = [("preingreso", "Preingreso"), ("ingreso", "Ingreso efectivo"), ("egreso", "Egreso")]

    # --- Para tablas exportables: Puedes guardar los querysets principales en context y usarlos para export a Excel/PDF ---

    context = {
        # KPIs GLOBAL (calculados sobre fichas_base)
        "total_fichas": total_fichas,
        "total_preingresos": total_preingresos,
        "total_ingresos": total_ingresos,
        "porcentaje_ingreso_efectivo": porcentaje_ingreso_efectivo,
        "total_egresos": total_egresos,
        "total_intervenciones": total_intervenciones,
        "casos_abiertos": casos_abiertos,
        "casos_cerrados": casos_cerrados,
        "promedio_duracion": promedio_duracion,  # calculado sobre fichas_base con egreso/ingreso

        # BREAKDOWNS Y GRÁFICOS (calculados sobre fichas filtradas)
        "genero_labels": genero_labels,
        "genero_data": genero_data,
        "rango_labels": rango_labels,
        "rango_data": rango_data,
        "nacionalidad_labels": nacionalidad_labels,
        "nacionalidad_data": nacionalidad_data,
        "prevision_labels": prevision_labels,
        "prevision_data": prevision_data,
        "ocupacion_labels": ocupacion_labels,
        "ocupacion_data": ocupacion_data,
        "barrio_labels": barrio_labels,
        "barrio_data": barrio_data,
        "cuadrante_labels": cuadrante_labels,
        "cuadrante_data": cuadrante_data,
        "motivo_labels": motivo_labels,
        "motivo_data": motivo_data,
        "profesional_labels": profesional_labels,
        "profesional_data": profesional_data,
        "diagnostico_labels": diagnostico_labels,
        "diagnostico_data": diagnostico_data,

        # DERIVACIONES (según fichas filtradas)
        "derivaciones": derivaciones,

        # EVOLUCIÓN TEMPORAL
        "fichas_mes_labels": fichas_mes_labels,
        "fichas_mes_data": fichas_mes_data,
        "egresos_mes_labels": egresos_mes_labels,
        "egresos_mes_data": egresos_mes_data,
        "intervenciones_mes_labels": intervenciones_mes_labels,
        "intervenciones_mes_data": intervenciones_mes_data,

        # FILTROS PARA SELECTS
        "barrios": barrios,
        "cuadrantes": cuadrantes,
        "motivos": motivos,
        "generos": generos,
        "profesionales": profesionales,
        "estados": estados,
        "etapas": etapas,

        # FILTROS ACTIVOS (para mostrar en selects)
        "filtros": {
            "barrio": barrio_id,
            "cuadrante": cuadrante_id,
            "genero": genero,
            "motivo": motivo,
            "profesional": profesional_id,
            "estado": estado,
            "etapa": etapa,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
        },

        # OPCIONAL: puedes pasar ambos querysets para debug o exportaciones
        "fichas_base": fichas_base,
        "fichas_filtradas": fichas,
        "total_derivaciones": total_derivaciones,

    }

    
    return render(request, "fichas/dashboard.html", context)

import pandas as pd
@login_required
def dashboard_export_excel(request):
    GET = request.GET
    barrio_id = GET.get('barrio')
    cuadrante_id = GET.get('cuadrante')
    genero = GET.get('genero')
    motivo = GET.get('motivo_ingreso')
    profesional_id = GET.get('profesional')
    estado = GET.get('estado')
    etapa = GET.get('etapa')
    fecha_inicio = GET.get('fecha_inicio')
    fecha_fin = GET.get('fecha_fin')

    fichas = FichaAcogida.objects.select_related('personaatendida', 'motivoingreso', 'egreso', 'profesional')
    if barrio_id:
        fichas = fichas.filter(personaatendida__barrio_id=barrio_id)
    if cuadrante_id:
        fichas = fichas.filter(personaatendida__cuadrante_id=cuadrante_id)
    if genero:
        fichas = fichas.filter(personaatendida__genero=genero)
    if motivo:
        fichas = fichas.filter(motivoingreso__motivo_ingreso=motivo)
    if profesional_id:
        fichas = fichas.filter(profesional_id=profesional_id)
    if fecha_inicio:
        fichas = fichas.filter(fecha_recepcion__gte=fecha_inicio)
    if fecha_fin:
        fichas = fichas.filter(fecha_recepcion__lte=fecha_fin)
    if estado == 'ABIERTA':
        fichas = fichas.filter(egreso__fecha_egreso__isnull=True)
    elif estado == 'CERRADA':
        fichas = fichas.filter(egreso__fecha_egreso__isnull=False)
    if etapa == "preingreso":
        fichas = fichas.filter(ingreso_efectivo=False)
    elif etapa == "ingreso":
        fichas = fichas.filter(ingreso_efectivo=True)
    elif etapa == "egreso":
        fichas = fichas.filter(egreso__isnull=False)

    data = []
    for f in fichas:
        persona = getattr(f, 'personaatendida', None)
        egreso = getattr(f, 'egreso', None)
        plan = getattr(f, 'planaccion', None)
        motivo = getattr(f, 'motivoingreso', None)

        # Para contar cuántas derivaciones existen
        num_derivaciones = 0
        if DerivacionSalud.objects.filter(ficha=f).exists(): num_derivaciones += 1
        if DerivacionCavd.objects.filter(ficha=f).exists(): num_derivaciones += 1
        if DerivacionUravit.objects.filter(ficha=f).exists(): num_derivaciones += 1
        if DerivacionOfam.objects.filter(ficha=f).exists(): num_derivaciones += 1
        if DerivacionDideco.objects.filter(ficha=f).exists(): num_derivaciones += 1
        if DerivacionGestionTerritorial.objects.filter(ficha=f).exists(): num_derivaciones += 1
        if DerivacionCapsUdla.objects.filter(ficha=f).exists(): num_derivaciones += 1
        if DerivacionOln.objects.filter(ficha=f).exists(): num_derivaciones += 1
        if DerivacionCdmCai.objects.filter(ficha=f).exists(): num_derivaciones += 1
        if DerivacionOtro.objects.filter(ficha=f).exists(): num_derivaciones += 1

        data.append({
            'ID': f.id,
            'Fecha recepción': f.fecha_recepcion,
            'Barrio': persona.barrio.nombre if persona and persona.barrio else 'No info',
            'Cuadrante': persona.cuadrante.nombre if persona and persona.cuadrante else 'No info',
            'Nombre': persona.nombre if persona else 'No info',
            'Género': persona.get_genero_display() if persona and persona.genero else 'No info',
            'Rango Etáreo': persona.get_rango_etareo_display() if persona and persona.rango_etareo else 'No info',
            'Nacionalidad': persona.nacionalidad if persona else 'No info',
            'Previsión': persona.get_prevision_display() if persona and persona.prevision else 'No info',
            'Estado': 'Cerrada' if egreso and egreso.fecha_egreso else 'Abierta',
            'Ingreso efectivo': 'Sí' if f.ingreso_efectivo else 'No',
            'Fecha ingreso': f.fecha_ingreso if f.fecha_ingreso else 'No info',
            'Fecha egreso': egreso.fecha_egreso if egreso and egreso.fecha_egreso else 'No info',
            'Motivo ingreso': motivo.get_motivo_ingreso_display() if motivo and motivo.motivo_ingreso else 'No info',
            'Número de intervenciones': f.intervencion_set.count(),
            'Número de derivaciones': num_derivaciones,
        })



    df = pd.DataFrame(data)
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="dashboard_pravim.xlsx"'
    df.to_excel(response, index=False)
    return response




from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse

@csrf_exempt
@login_required
def api_intervencion_eliminar(request, intervencion_id):
    if request.method == "POST":
        try:
            from .models import Intervencion
            interv = Intervencion.objects.get(id=intervencion_id)
            # REGISTRAR EN BITÁCORA
            registrar_cambios_bitacora(interv, None, request.user, interv.ficha, accion="ELIMINADO")   
            interv.delete()
            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Método no permitido"})


from .models import (
    FichaAcogida, PersonaAtendida, DerivacionCavd, DerivacionUravit, DerivacionCdmCai,
    DerivacionSalud, DerivacionOfam, DerivacionDideco, DerivacionGestionTerritorial,
    DerivacionCapsUdla, DerivacionOln, DerivacionOtro
)
from django.db.models import Q

DERIVACION_MODELS = [
    ('CAVD', DerivacionCavd, 'fecha_derivacion_cavd', 'fecha_respuesta_cavd', 'fecha_ingreso_cavd'),
    ('URAVIT', DerivacionUravit, 'fecha_derivacion_uravit', 'fecha_respuesta_uravit', None),
    ('CDM-CAI', DerivacionCdmCai, 'fecha_derivacion_cdm_cai', 'fecha_respuesta_cdm_cai', 'fecha_ingreso_cdm_cai'),
    ('SALUD', DerivacionSalud, 'fecha_derivacion_salud', 'fecha_respuesta_salud', 'fecha_ingreso_salud'),
    ('OFAM', DerivacionOfam, 'fecha_derivacion_ofam', 'fecha_respuesta_ofam', 'fecha_ingreso_ofam'),
    ('DIDECO', DerivacionDideco, 'fecha_derivacion_dideco', 'fecha_respuesta_dideco', 'fecha_ingreso_dideco'),
    ('GESTIÓN TERRITORIAL', DerivacionGestionTerritorial, 'fecha_derivacion_gestion_territorial', 'fecha_respuesta_gestion_territorial', 'fecha_ingreso_gestion_territorial'),
    ('CAPS UDLA', DerivacionCapsUdla, 'fecha_derivacion_caps_udla', 'fecha_respuesta_caps_udla', 'fecha_ingreso_caps_udla'),
    ('OLN', DerivacionOln, 'fecha_derivacion_oln', 'fecha_respuesta_oln', 'fecha_ingreso_oln'),
    ('OTRO', DerivacionOtro, 'fecha_derivacion_otro', 'fecha_respuesta_otro', 'fecha_ingreso_otro'),
]
from django.db.models import Q
from datetime import datetime, date
from .models import FichaAcogida, PersonaAtendida
from .forms import BuscarDerivacionForm  # lo crearás más abajo

from datetime import datetime

from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.shortcuts import render

@login_required
def buscar_derivacion(request):
    derivaciones = []

    for nombre, modelo, campo_der, campo_resp, campo_ing in DERIVACION_MODELS:
        for obj in modelo.objects.select_related('ficha__personaatendida').all():
            derivaciones.append({
                'nombre': nombre,
                'ficha': obj.ficha,
                'persona': obj.ficha.personaatendida,
                'fecha_recepcion': obj.ficha.fecha_recepcion,
                'fecha_derivacion': getattr(obj, campo_der, None),
                'fecha_respuesta': getattr(obj, campo_resp, None),
                'fecha_ingreso': getattr(obj, campo_ing, None) if campo_ing else None,
            })

    form = BuscarDerivacionForm(request.GET or None)
    if form.is_valid():
        cd = form.cleaned_data
        numero_ficha = request.GET.get('ficha')
        if numero_ficha:
            derivaciones = [
                d for d in derivaciones
                if d['ficha'] and str(d['ficha'].id) == str(numero_ficha)
            ]
        if cd.get('nombre'):
            derivaciones = [
                d for d in derivaciones
                if d['persona'].nombre and cd['nombre'].lower() in d['persona'].nombre.lower()
            ]

        if cd.get('rut'):
            derivaciones = [
                d for d in derivaciones
                if d['persona'].rut_pasaporte and cd['rut'].lower() in d['persona'].rut_pasaporte.lower()
            ]

        if cd.get('tipo_derivacion'):
            derivaciones = [
                d for d in derivaciones
                if cd['tipo_derivacion'].lower() in d['nombre'].lower()
            ]

        if cd.get('fecha_inicio'):
            derivaciones = [
                d for d in derivaciones
                if d['fecha_derivacion'] and d['fecha_derivacion'] >= cd['fecha_inicio']
            ]

        if cd.get('fecha_fin'):
            derivaciones = [
                d for d in derivaciones
                if d['fecha_derivacion'] and d['fecha_derivacion'] <= cd['fecha_fin']
            ]

        if cd.get('solo_con_ingreso'):
            derivaciones = [
                d for d in derivaciones
                if d['fecha_ingreso'] is not None
            ]
        if cd.get('solo_con_derivacion'):
            derivaciones = [
                d for d in derivaciones
                if d['fecha_derivacion'] is not None
            ]
        if cd.get('solo_con_recepcion'):
            derivaciones = [
                d for d in derivaciones
                if d['ficha'].fecha_recepcion is not None
            ]

        if cd.get('solo_con_respuesta'):
            derivaciones = [
                d for d in derivaciones
                if d['fecha_respuesta'] is not None
            ]

    # Ordenar por N° ficha (id) descendente
    derivaciones = sorted(derivaciones, key=lambda x: x['ficha'].id if x['ficha'] else 0, reverse=True)

    # Filtros por fechas vacías
    if request.GET.get('sin_derivacion'):
        derivaciones = [d for d in derivaciones if not d['fecha_derivacion']]

    if request.GET.get('sin_respuesta'):
        derivaciones = [d for d in derivaciones if not d['fecha_respuesta']]

    if request.GET.get('sin_ingreso'):
        derivaciones = [d for d in derivaciones if not d['fecha_ingreso']]

    # ---- PAGINADOR ----
    paginator = Paginator(derivaciones, 30)  # 30 por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'fichas/buscar_derivacion.html', {
        'form': form,
        'page_obj': page_obj,   # Para iterar sobre page_obj en el template
        'request': request,     # Para pasar los filtros actuales y mantenerlos en la paginación
    })



from django.shortcuts import get_object_or_404
from django.template.loader import get_template
from django.http import HttpResponse
from weasyprint import HTML, CSS
from datetime import datetime
from .models import (
    FichaAcogida, PersonaAtendida, Denuncia, MotivoIngreso,
    AspectosPsicologicos, RedesApoyo, PlanAccion, Egreso, Intervencion,
    DerivacionCavd, DerivacionUravit, DerivacionCdmCai, DerivacionSalud,
    DerivacionOfam, DerivacionDideco, DerivacionGestionTerritorial,
    DerivacionCapsUdla, DerivacionOln, DerivacionOtro
)
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from weasyprint import CSS

custom_css = CSS(string="""
    @page {
        size: A4;
        margin: 80pt 28pt 95pt 28pt; /* Arriba, derecha, abajo, izquierda (en puntos, no px) */
        @top-center { content: element(header); }
        @bottom-center { content: element(footer); }
        @bottom-right { content: "Página " counter(page) " de " counter(pages); }
    }
    .header { position: running(header); }
    .footer { position: running(footer); }
    tr, td, th, .card { page-break-inside: avoid; break-inside: avoid; }
    body { font-family: Arial, sans-serif; }
""")


@login_required
def ficha_pdf(request, ficha_id):
    ficha = get_object_or_404(FichaAcogida, id=ficha_id)
    persona = getattr(ficha, "personaatendida", None)
    denuncia = getattr(ficha, "denuncia", None)
    motivo = getattr(ficha, "motivoingreso", None)
    psicologico = getattr(ficha, "aspectospsicologicos", None)
    redes = getattr(ficha, "redesapoyo", None)
    plan = getattr(ficha, "planaccion", None)
    egreso = getattr(ficha, "egreso", None)
    intervenciones = Intervencion.objects.filter(ficha=ficha).order_by('fecha')

    # Derivaciones usando nombres de modelos correctos
    deriv_cavd = DerivacionCavd.objects.filter(ficha=ficha).first()
    deriv_uravit = DerivacionUravit.objects.filter(ficha=ficha).first()
    deriv_cdm_cai = DerivacionCdmCai.objects.filter(ficha=ficha).first()
    deriv_salud = DerivacionSalud.objects.filter(ficha=ficha).first()
    deriv_ofam = DerivacionOfam.objects.filter(ficha=ficha).first()
    deriv_dideco = DerivacionDideco.objects.filter(ficha=ficha).first()
    deriv_gestion_territorial = DerivacionGestionTerritorial.objects.filter(ficha=ficha).first()
    deriv_caps_udla = DerivacionCapsUdla.objects.filter(ficha=ficha).first()
    deriv_oln = DerivacionOln.objects.filter(ficha=ficha).first()
    deriv_otro = DerivacionOtro.objects.filter(ficha=ficha).first()

    context = {
        'ficha': ficha,
        'persona': persona,
        'denuncia': denuncia,
        'motivo': motivo,
        'psicologico': psicologico,
        'redes': redes,
        'plan': plan,
        'egreso': egreso,
        'intervenciones': intervenciones,
        # Derivaciones:
        'deriv_cavd': deriv_cavd,
        'deriv_uravit': deriv_uravit,
        'deriv_cdm_cai': deriv_cdm_cai,
        'deriv_salud': deriv_salud,
        'deriv_ofam': deriv_ofam,
        'deriv_dideco': deriv_dideco,
        'deriv_gestion_territorial': deriv_gestion_territorial,
        'deriv_caps_udla': deriv_caps_udla,
        'deriv_oln': deriv_oln,
        'deriv_otro': deriv_otro,
        # Extras para PDF:
        'now': datetime.now(),
        'user': request.user,
    }

    template = get_template('fichas/ficha_pdf.html')
    html_string = template.render(context, request=request)
    base_url = request.build_absolute_uri('/')

    # CSS especial para WeasyPrint (puedes personalizar si quieres encabezados/pies por página)
    custom_css = CSS(string="""
        @page {
            size: A4;
            margin: 42pt 24pt 66pt 24pt;
            @top-center {
                content: element(header)
            }
            @bottom-center {
                content: element(footer)
            }
        }
        .header { position: running(header); }
        .footer { position: running(footer); }
        tr, td, th, .card { page-break-inside: avoid; break-inside: avoid; }
        body { font-family: Arial, sans-serif; }
    """)

    pdf_file = HTML(string=html_string, base_url=base_url).write_pdf(stylesheets=[custom_css])

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Ficha_{ficha.id}.pdf"'
    return response


from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from datetime import datetime
from .models import (
    FichaAcogida, PersonaAtendida, Denuncia, MotivoIngreso,
    AspectosPsicologicos, RedesApoyo, PlanAccion, Egreso, Intervencion,
    DerivacionCavd, DerivacionUravit, DerivacionCdmCai, DerivacionSalud,
    DerivacionOfam, DerivacionDideco, DerivacionGestionTerritorial,
    DerivacionCapsUdla, DerivacionOln, DerivacionOtro
)

@login_required
def ficha_pdf_preview(request, ficha_id):
    ficha = get_object_or_404(FichaAcogida, id=ficha_id)
    persona = getattr(ficha, "personaatendida", None)
    denuncia = getattr(ficha, "denuncia", None)
    motivo = getattr(ficha, "motivoingreso", None)
    psicologico = getattr(ficha, "aspectospsicologicos", None)
    redes = getattr(ficha, "redesapoyo", None)
    plan = getattr(ficha, "planaccion", None)
    egreso = getattr(ficha, "egreso", None)
    intervenciones = Intervencion.objects.filter(ficha=ficha).order_by('fecha')

    deriv_cavd = DerivacionCavd.objects.filter(ficha=ficha).first()
    deriv_uravit = DerivacionUravit.objects.filter(ficha=ficha).first()
    deriv_cdm_cai = DerivacionCdmCai.objects.filter(ficha=ficha).first()
    deriv_salud = DerivacionSalud.objects.filter(ficha=ficha).first()
    deriv_ofam = DerivacionOfam.objects.filter(ficha=ficha).first()
    deriv_dideco = DerivacionDideco.objects.filter(ficha=ficha).first()
    deriv_gestion_territorial = DerivacionGestionTerritorial.objects.filter(ficha=ficha).first()
    deriv_caps_udla = DerivacionCapsUdla.objects.filter(ficha=ficha).first()
    deriv_oln = DerivacionOln.objects.filter(ficha=ficha).first()
    deriv_otro = DerivacionOtro.objects.filter(ficha=ficha).first()

    context = {
        'ficha': ficha,
        'persona': persona,
        'denuncia': denuncia,
        'motivo': motivo,
        'psicologico': psicologico,
        'redes': redes,
        'plan': plan,
        'egreso': egreso,
        'intervenciones': intervenciones,
        'deriv_cavd': deriv_cavd,
        'deriv_uravit': deriv_uravit,
        'deriv_cdm_cai': deriv_cdm_cai,
        'deriv_salud': deriv_salud,
        'deriv_ofam': deriv_ofam,
        'deriv_dideco': deriv_dideco,
        'deriv_gestion_territorial': deriv_gestion_territorial,
        'deriv_caps_udla': deriv_caps_udla,
        'deriv_oln': deriv_oln,
        'deriv_otro': deriv_otro,
        'now': timezone.now(),
        'user': request.user,
    }
    return render(request, 'fichas/ficha_pdf.html', context)



@login_required
@login_required
def historial_cambios(request):
    cambios = FichaChangeLog.objects.select_related('ficha', 'usuario').all()
    # FILTROS
    numero_ficha = request.GET.get('numero_ficha')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    if numero_ficha:
        cambios = cambios.filter(ficha__id=numero_ficha)
    if fecha_inicio:
        cambios = cambios.filter(fecha__date__gte=fecha_inicio)
    if fecha_fin:
        cambios = cambios.filter(fecha__date__lte=fecha_fin)

    cambios = cambios.order_by('-fecha')  # Más recientes arriba

    paginator = Paginator(cambios, 30)  # 30 por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'fichas/historial_cambios.html', {
        'page_obj': page_obj,
        'request': request,
    })
    

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.utils.encoding import smart_str

@login_required
def historial_ficha(request, ficha_id):
    ficha = get_object_or_404(FichaAcogida, id=ficha_id)
    cambios = ficha.changelog.select_related('usuario').all()
    return render(request, 'fichas/historial_ficha.html', {'ficha': ficha, 'cambios': cambios})


def descargar_historial_excel(request):
    # 1. Obtén los filtros GET (ajusta los nombres según tu form)
    numero_ficha = request.GET.get('numero_ficha')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')

    # 2. Aplica los filtros al queryset
    historial = FichaChangeLog.objects.all()
    if numero_ficha:
        historial = historial.filter(ficha__id=numero_ficha)
    if fecha_inicio:
        historial = historial.filter(fecha__date__gte=fecha_inicio)
    if fecha_fin:
        historial = historial.filter(fecha__date__lte=fecha_fin)

    # 3. Ordena por fecha descendente (opcional)
    historial = historial.order_by('-fecha')

    # 4. Crea el Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial Cambios"

    # 5. Escribe el encabezado
    headers = [
        "Fecha", "Ficha", "Usuario", "Acción", "Campo",
        "Valor anterior", "Valor nuevo"
    ]
    ws.append(headers)

    # 6. Escribe los datos
    for h in historial:
        ws.append([
            h.fecha.strftime('%d/%m/%Y %H:%M'),
            f'#{h.ficha.id}' if h.ficha else '',
            h.usuario.get_full_name() if h.usuario.get_full_name() else h.usuario.username,
            h.get_accion_display(),
            h.campo.title() if h.campo else '',
            h.valor_anterior if h.valor_anterior else '—',
            h.valor_nuevo if h.valor_nuevo else '—'
        ])

    # Ajusta el ancho de columnas
    for col_num, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    # 7. Prepara la respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = "historial_cambios.xlsx"
    response['Content-Disposition'] = f'attachment; filename={smart_str(filename)}'

    wb.save(response)
    return response

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.utils.encoding import smart_str
from .models import (
    DerivacionCavd, DerivacionUravit, DerivacionCdmCai, DerivacionSalud, DerivacionOfam, 
    DerivacionDideco, DerivacionGestionTerritorial, DerivacionCapsUdla, DerivacionOln, DerivacionOtro
)

DERIVACION_MODELS = [
    ('CAVD', DerivacionCavd, 'fecha_derivacion_cavd', 'fecha_respuesta_cavd', 'fecha_ingreso_cavd'),
    ('URAVIT', DerivacionUravit, 'fecha_derivacion_uravit', 'fecha_respuesta_uravit', None),
    ('CDM-CAI', DerivacionCdmCai, 'fecha_derivacion_cdm_cai', 'fecha_respuesta_cdm_cai', 'fecha_ingreso_cdm_cai'),
    ('SALUD', DerivacionSalud, 'fecha_derivacion_salud', 'fecha_respuesta_salud', 'fecha_ingreso_salud'),
    ('OFAM', DerivacionOfam, 'fecha_derivacion_ofam', 'fecha_respuesta_ofam', 'fecha_ingreso_ofam'),
    ('DIDECO', DerivacionDideco, 'fecha_derivacion_dideco', 'fecha_respuesta_dideco', 'fecha_ingreso_dideco'),
    ('GESTIÓN TERRITORIAL', DerivacionGestionTerritorial, 'fecha_derivacion_gestion_territorial', 'fecha_respuesta_gestion_territorial', 'fecha_ingreso_gestion_territorial'),
    ('CAPS UDLA', DerivacionCapsUdla, 'fecha_derivacion_caps_udla', 'fecha_respuesta_caps_udla', 'fecha_ingreso_caps_udla'),
    ('OLN', DerivacionOln, 'fecha_derivacion_oln', 'fecha_respuesta_oln', 'fecha_ingreso_oln'),
    ('OTRO', DerivacionOtro, 'fecha_derivacion_otro', 'fecha_respuesta_otro', 'fecha_ingreso_otro'),
]

def descargar_derivacion_excel(request):
    # Filtros
    nombre = request.GET.get('nombre', '').strip()
    rut = request.GET.get('rut', '').strip()
    tipo_derivacion = request.GET.get('tipo_derivacion', '')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    sin_derivacion = request.GET.get('sin_derivacion')
    sin_respuesta = request.GET.get('sin_respuesta')
    sin_ingreso = request.GET.get('sin_ingreso')
    numero_ficha = request.GET.get('ficha')

    # Juntar todas las derivaciones de todos los modelos
    derivaciones = []
    for nombre_tipo, modelo, campo_der, campo_resp, campo_ing in DERIVACION_MODELS:
        for obj in modelo.objects.select_related('ficha__personaatendida').all():
            derivaciones.append({
                'nombre': nombre_tipo,
                'ficha': obj.ficha,
                'persona': getattr(obj.ficha, 'personaatendida', None),
                'fecha_recepcion': obj.ficha.fecha_recepcion,
                'fecha_derivacion': getattr(obj, campo_der, None),
                'fecha_respuesta': getattr(obj, campo_resp, None),
                'fecha_ingreso': getattr(obj, campo_ing, None) if campo_ing else None,
            })

    # Aplicar los mismos filtros que en buscar_derivacion
    if numero_ficha:
        derivaciones = [
            d for d in derivaciones
            if d['ficha'] and str(d['ficha'].id) == str(numero_ficha)
        ]

    if nombre:
        derivaciones = [
            d for d in derivaciones
            if d['persona'] and d['persona'].nombre and nombre.lower() in d['persona'].nombre.lower()
        ]

    if rut:
        derivaciones = [
            d for d in derivaciones
            if d['persona'] and d['persona'].rut_pasaporte and rut.lower() in d['persona'].rut_pasaporte.lower()
        ]

    if tipo_derivacion:
        derivaciones = [
            d for d in derivaciones
            if tipo_derivacion.lower() in d['nombre'].lower()
        ]

    if fecha_inicio:
        derivaciones = [
            d for d in derivaciones
            if d['fecha_derivacion'] and d['fecha_derivacion'] >= fecha_inicio
        ]
    if fecha_fin:
        derivaciones = [
            d for d in derivaciones
            if d['fecha_derivacion'] and d['fecha_derivacion'] <= fecha_fin
        ]

    if sin_derivacion:
        derivaciones = [d for d in derivaciones if not d['fecha_derivacion']]
    if sin_respuesta:
        derivaciones = [d for d in derivaciones if not d['fecha_respuesta']]
    if sin_ingreso:
        derivaciones = [d for d in derivaciones if not d['fecha_ingreso']]

    # Ordenar por N° ficha descendente
    derivaciones = sorted(derivaciones, key=lambda x: x['ficha'].id if x['ficha'] else 0, reverse=True)

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Derivaciones"

    headers = [
        "N° Ficha", "Fecha Recepción", "Nombre Persona", "RUT", "Tipo de Derivación",
        "Fecha Derivación", "Fecha Respuesta", "Fecha Ingreso"
    ]
    ws.append(headers)

    for d in derivaciones:
        persona = d['persona']
        ws.append([
            d['ficha'].id if d['ficha'] else '',
            d['fecha_recepcion'].strftime('%d/%m/%Y') if d['fecha_recepcion'] else '',
            persona.nombre if persona else '-',
            persona.rut_pasaporte if persona else '-',
            d['nombre'],
            d['fecha_derivacion'].strftime('%d/%m/%Y') if d['fecha_derivacion'] else '-',
            d['fecha_respuesta'].strftime('%d/%m/%Y') if d['fecha_respuesta'] else '-',
            d['fecha_ingreso'].strftime('%d/%m/%Y') if d['fecha_ingreso'] else '-',
        ])

    # Ajustar ancho de columnas
    for col_num, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = "derivaciones.xlsx"
    response['Content-Disposition'] = f'attachment; filename={smart_str(filename)}'
    wb.save(response)
    return response


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.utils.encoding import smart_str
from .models import FichaAcogida

@login_required
def buscar_persona_excel(request):
    # Mismos filtros que en buscar_persona
    fichas = FichaAcogida.objects.all().select_related('personaatendida', 'motivoingreso', 'egreso')
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    numero_ficha = request.GET.get('numero_ficha')
    nombre = request.GET.get('nombre')
    rut = request.GET.get('rut')
    domicilio = request.GET.get('domicilio')
    delito = request.GET.get('delito')
    estado = request.GET.get('estado')
    ingreso = request.GET.get('ingreso')

    if ingreso == "con_ingreso":
        fichas = fichas.filter(ingreso_efectivo=True)
    elif ingreso == "sin_ingreso":
        fichas = fichas.filter(ingreso_efectivo=False)

    if fecha_inicio:
        fichas = fichas.filter(fecha_recepcion__gte=fecha_inicio)
    if fecha_fin:
        fichas = fichas.filter(fecha_recepcion__lte=fecha_fin)
    if numero_ficha:
        fichas = fichas.filter(id=numero_ficha)
    if nombre:
        fichas = fichas.filter(personaatendida__nombre__icontains=nombre)
    if delito:
        fichas = fichas.filter(motivoingreso__motivo_ingreso=delito)
    if estado == "ABIERTA":
        fichas = fichas.filter(egreso__isnull=True)
    elif estado == "CERRADA":
        fichas = fichas.filter(egreso__isnull=False)
    if rut:
        fichas = fichas.filter(personaatendida__rut_pasaporte__icontains=rut)
    if domicilio:
        fichas = fichas.filter(personaatendida__direccion__icontains=domicilio)

    # Arma el Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Personas Atendidas"

    headers = [
        "N° Ficha", "Fecha Recepción", "Nombre", "RUT", "Domicilio", "Motivo/Delito",
        "Estado", "Ingreso Efectivo", "Fecha Ingreso", "Fecha Egreso", "Últ. intervención"
    ]
    ws.append(headers)

    for ficha in fichas:
        persona = getattr(ficha, 'personaatendida', None)
        egreso = getattr(ficha, 'egreso', None)
        motivo = getattr(ficha, 'motivoingreso', None)
        ultima = ficha.intervencion_set.order_by('-fecha').first()

        ws.append([
            ficha.id,
            ficha.fecha_recepcion.strftime('%d/%m/%Y') if ficha.fecha_recepcion else '',
            persona.nombre if persona else '',
            persona.rut_pasaporte if persona else '',
            persona.direccion if persona else '',
            motivo.get_motivo_ingreso_display() if motivo and motivo.motivo_ingreso else '',
            "Cerrada" if egreso and egreso.fecha_egreso else "Abierta",
            "Sí" if ficha.ingreso_efectivo else "No",
            ficha.fecha_ingreso.strftime('%d/%m/%Y') if ficha.fecha_ingreso else '',
            egreso.fecha_egreso.strftime('%d/%m/%Y') if egreso and egreso.fecha_egreso else '',
            ultima.fecha.strftime('%d/%m/%Y') if ultima else ''
        ])

    # Ajustar ancho columnas
    for col_num, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = "personas_atendidas.xlsx"
    response['Content-Disposition'] = f'attachment; filename={smart_str(filename)}'
    wb.save(response)
    return response
